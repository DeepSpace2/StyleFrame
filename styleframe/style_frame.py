import datetime as dt
import numpy as np
import pathlib
import pandas as pd

from functools import partial

from .deprecations import deprecated_kwargs
from . import utils
from copy import deepcopy
from collections import OrderedDict
from collections.abc import Iterable
from openpyxl import load_workbook
from openpyxl.cell.cell import get_column_letter
from openpyxl.xml.functions import fromstring, QName
from openpyxl.utils import cell

from styleframe.container import Container
from styleframe.series import Series
from styleframe.styler import Styler, ColorScaleConditionalFormatRule

try:
    pd_timestamp = pd.Timestamp
except AttributeError:
    pd_timestamp = pd.tslib.Timestamp


class StyleFrame:
    """
    A wrapper class that wraps a :class:`pandas.DataFrame` object and represent a stylized dataframe.
    Stores container objects that have values and styles that will be applied to excel

    :param obj: Any object that pandas' dataframe can be initialized with: an existing dataframe, a dictionary,
          a list of dictionaries or another StyleFrame.
    :param styler_obj: Will be used as the default style of all cells.
    :type styler_obj: :class:`.Styler`
    """
    P_FACTOR = 1.3
    A_FACTOR = 13

    def __init__(self, obj, styler_obj=None):
        from_another_styleframe = False
        from_pandas_dataframe = False
        if styler_obj and not isinstance(styler_obj, Styler):
            raise TypeError('styler_obj must be {}, got {} instead.'.format(Styler.__name__, type(styler_obj).__name__))
        if isinstance(obj, pd.DataFrame):
            from_pandas_dataframe = True
            if obj.empty:
                self.data_df = deepcopy(obj)
            else:
                self.data_df = obj.applymap(lambda x: Container(x, deepcopy(styler_obj)) if not isinstance(x, Container) else x)
        elif isinstance(obj, pd.Series):
            self.data_df = obj.apply(lambda x: Container(x, deepcopy(styler_obj)) if not isinstance(x, Container) else x)
        elif isinstance(obj, (dict, list)):
            self.data_df = pd.DataFrame(obj).applymap(lambda x: Container(x, deepcopy(styler_obj)) if not isinstance(x, Container) else x)
        elif isinstance(obj, StyleFrame):
            self.data_df = deepcopy(obj.data_df)
            from_another_styleframe = True
        else:
            raise TypeError("{} __init__ doesn't support {}".format(type(self).__name__, type(obj).__name__))
        self.data_df.columns = [Container(col, deepcopy(styler_obj)) if not isinstance(col, Container) else deepcopy(col)
                                for col in self.data_df.columns]
        self.data_df.index = [Container(index, deepcopy(styler_obj)) if not isinstance(index, Container) else deepcopy(index)
                              for index in self.data_df.index]

        if from_pandas_dataframe:
            self.data_df.index.name = obj.index.name

        self._columns_width = obj._columns_width if from_another_styleframe else OrderedDict()
        self._rows_height = obj._rows_height if from_another_styleframe else OrderedDict()
        self._has_custom_headers_style = obj._has_custom_headers_style if from_another_styleframe else False
        self._cond_formatting = []
        self._default_style = styler_obj or Styler()
        self._index_header_style = obj._index_header_style if from_another_styleframe else self._default_style

        self._known_attrs = {'at': self.data_df.at,
                             'loc': self.data_df.loc,
                             'iloc': self.data_df.iloc,
                             'applymap': self.data_df.applymap,
                             'groupby': self.data_df.groupby,
                             'index': self.data_df.index,
                             'fillna': self.data_df.fillna}

    def __str__(self):
        return str(self.data_df)

    def __len__(self):
        return len(self.data_df)

    def __getitem__(self, item):
        if isinstance(item, pd.Series):
            return self.data_df.__getitem__(item).index
        if isinstance(item, list):
            return StyleFrame(self.data_df.__getitem__(item))
        return Series(self.data_df.__getitem__(item))

    def __setitem__(self, key, value):
        if isinstance(value, (Iterable, pd.Series)):
            self.data_df.__setitem__(Container(key), list(map(Container, value)))
        else:
            self.data_df.__setitem__(Container(key), Container(value))

    def __delitem__(self, item):
        return self.data_df.__delitem__(item)

    def __getattr__(self, attr):
        if attr in self.data_df.columns:
            return self.data_df[attr]
        try:
            return self._known_attrs[attr]
        except KeyError:
            raise AttributeError("'{}' object has no attribute '{}'".format(type(self).__name__, attr))

    @property
    def columns(self):
        return self.data_df.columns

    @columns.setter
    def columns(self, columns):
        self.data_df.columns = [col if isinstance(col, Container) else Container(value=col)
                                for col in columns]

    def _get_column_as_letter(self, sheet, column_to_convert, startcol=0):
        col = column_to_convert.value if isinstance(column_to_convert, Container) else column_to_convert
        if not isinstance(col, (int, str)):
            raise TypeError("column must be an index, column letter or column name")
        column_as_letter = None
        if col in self.data_df.columns:  # column name
            column_index = self.data_df.columns.get_loc(col) + startcol + 1  # worksheet columns index start from 1
            column_as_letter = cell.get_column_letter(column_index)

        # column index
        elif isinstance(col, int) and col >= 1:
            column_as_letter = cell.get_column_letter(startcol + col)

        # assuming we got column letter
        elif isinstance(col, str) and col <= get_column_letter(sheet.max_column):
            column_as_letter = col

        if column_as_letter is None or cell.column_index_from_string(column_as_letter) > sheet.max_column:
            raise IndexError("column: %s is out of columns range." % column_to_convert)
        return column_as_letter

    @classmethod
    def read_excel(cls, path, sheet_name=0, read_style=False, use_openpyxl_styles=False,
                   read_comments=False, **kwargs):
        """
        Creates a StyleFrame object from an existing Excel.

        .. note:: :meth:`read_excel` also accepts all arguments that :func:`pandas.read_excel` accepts as kwargs.

        :param str path: The path to the Excel file to read.
        :param sheetname:
              .. deprecated:: 1.6
                 Use ``sheet_name`` instead.
              .. versionchanged:: 4.0
                 Removed
        :param sheet_name: The sheet name to read. If an integer is provided then it be used as a zero-based
                sheet index. Default is 0.
        :type sheet_name: str or int
        :param bool read_style: If ``True`` the sheet's style will be loaded to the returned StyleFrame object.
        :param bool use_openpyxl_styles: If ``True`` (and `read_style` is also ``True``) then the styles in the returned
            StyleFrame object will be Openpyxl's style objects. If ``False``, the styles will be :class:`.Styler` objects.

            .. note:: Using ``use_openpyxl_styles=False`` is useful if you are going to filter columns or rows by style, for example:

                     ::

                        sf = sf[[col for col in sf.columns if col.style.font == utils.fonts.arial]]

        :param bool read_comments: If ``True`` (and `read_style` is also ``True``) cells' comments will be loaded to the returned StyleFrame object. Note
                that reading comments without reading styles is currently not supported.

        :return: StyleFrame object
        :rtype: :class:`StyleFrame`
        """

        def _get_scheme_colors_from_excel(wb):
            xlmns = 'http://schemas.openxmlformats.org/drawingml/2006/main'
            if wb.loaded_theme is None:
                return []
            root = fromstring(wb.loaded_theme)
            theme_element = root.find(QName(xlmns, 'themeElements').text)
            color_schemes = theme_element.findall(QName(xlmns, 'clrScheme').text)
            colors = []
            for colorScheme in color_schemes:
                for tag in ['lt1', 'dk1', 'lt2', 'dk2', 'accent1', 'accent2', 'accent3', 'accent4', 'accent5', 'accent6']:
                    accent = list(colorScheme.find(QName(xlmns, tag).text))[0]
                    if 'window' in accent.attrib['val']:
                        colors.append(accent.attrib['lastClr'])
                    else:
                        colors.append(accent.attrib['val'])
            return colors

        def _get_style_object(sheet, theme_colors, row, column):
            cell = sheet.cell(row=row, column=column)
            if use_openpyxl_styles:
                return cell
            else:
                return Styler.from_openpyxl_style(cell, theme_colors,
                                                  read_comments and cell.comment)

        def _read_style():
            wb = load_workbook(path)
            if isinstance(sheet_name, str):
                sheet = wb[sheet_name]
            elif isinstance(sheet_name, int):
                sheet = wb.worksheets[sheet_name]
            else:
                raise TypeError("'sheet_name' must be a string or int, got {} instead".format(type(sheet_name)))
            theme_colors = _get_scheme_colors_from_excel(wb)

            # Set the headers row height
            if header_arg is not None:
                headers_row_idx = header_arg + 1
                sf._rows_height[headers_row_idx] = sheet.row_dimensions[headers_row_idx].height

            get_style_object = partial(_get_style_object, sheet=sheet, theme_colors=theme_colors)
            for col_index, col_name in enumerate(sf.columns):
                col_index_in_excel = col_index + 1
                if col_index_in_excel == excel_index_col:
                    for row_index, sf_index in enumerate(sf.index, start=2):
                        sf_index.style = get_style_object(row=row_index, column=col_index_in_excel)
                    col_index_in_excel += 1  # Move next to excel indices column

                sf.columns[col_index].style = get_style_object(row=1, column=col_index_in_excel)
                for row_index, sf_index in enumerate(sf.index, start=start_row_index):
                    sf.at[sf_index, col_name].style = get_style_object(row=row_index, column=col_index_in_excel)
                    sf._rows_height[row_index] = sheet.row_dimensions[row_index].height

                sf._columns_width[col_name] = sheet.column_dimensions[sf._get_column_as_letter(sheet, col_name)].width

        header_arg = kwargs.get('header', 0)
        if read_style and isinstance(header_arg, Iterable):
            raise ValueError('Not supporting multiple index columns with read style.')
        if header_arg is None:
            start_row_index = 1
        else:
            start_row_index = header_arg + 2
        index_col = kwargs.get('index_col')
        excel_index_col = index_col + 1 if index_col is not None else None
        if read_style and isinstance(excel_index_col, Iterable):
            raise ValueError('Not supporting multiple index columns with read style.')

        sf = cls(pd.read_excel(path, sheet_name, **kwargs))
        if read_style:
            _read_style()
            sf._has_custom_headers_style = True

        return sf

    @classmethod
    def read_excel_as_template(cls, path, df, use_df_boundaries=False, **kwargs):
        """
        .. versionadded:: 3.0.1

        Create a StyleFrame object from an excel template with data of the given DataFrame.

        .. note:: :meth:`read_excel_as_template` also accepts all arguments that :meth:`read_excel` accepts as kwargs except for ``read_style`` which must be ``True``.

        :param str path: The path to the Excel file to read.
        :param df: The data to apply to the given template.
        :type df: :class:`pandas.DataFrame`
        :param bool use_df_boundaries: If ``True`` the template will be cut according to the boundaries of the given DataFrame.

        :return: StyleFrame object
        :rtype: :class:`StyleFrame`
        """

        sf = cls.read_excel(path=path, read_style=True, **kwargs)

        num_of_rows, num_of_cols = len(df.index), len(df.columns)
        template_num_of_rows, template_num_of_cols = len(sf.index), len(sf.columns)

        num_of_cols_to_copy_with_style = min(num_of_cols, template_num_of_cols)
        num_of_rows_to_copy_with_style = min(num_of_rows, template_num_of_rows)
        for col_index in range(num_of_cols_to_copy_with_style):
            for row_index in range(num_of_rows_to_copy_with_style):
                sf.iloc[row_index, col_index].value = df.iloc[row_index, col_index]

        # Insert extra data in cases where the df is larger than the template.
        for extra_col in df.columns[template_num_of_cols:]:
            sf[extra_col] = df[extra_col][:template_num_of_rows]
        for row_index in df.index[template_num_of_rows:]:
            sf_index = Container(value=row_index)
            sf.loc[sf_index] = list(map(Container, df.loc[row_index]))

        sf.rename({sf.columns[col_index].value: df_col
                   for col_index, df_col in enumerate(df.columns)},
                  inplace=True)

        if use_df_boundaries:
            sf.data_df = sf.data_df.iloc[:num_of_rows, :num_of_cols]
            rows_height = OrderedDict()
            rows_height_range = range(num_of_rows)
            for i, (k, v) in enumerate(sf._rows_height.items()):
                if i in rows_height_range:
                    rows_height[k] = v
            sf._rows_height = rows_height

            columns_width = OrderedDict()
            columns_width_range = range(num_of_cols)
            for i, (k, v) in enumerate(sf._columns_width.items()):
                if i in columns_width_range:
                    columns_width[k] = v
            sf._columns_width = columns_width
        return sf

    # noinspection PyPep8Naming
    @classmethod
    def ExcelWriter(cls, path, **kwargs):
        """
        A shortcut for :class:`pandas.ExcelWriter`, and accepts any argument it accepts except for ``engine``
        """

        if 'engine' in kwargs:
            raise ValueError('`engine` argument for StyleFrame.ExcelWriter can not be set')
        return pd.ExcelWriter(path, engine='openpyxl', **kwargs)

    @property
    def row_indexes(self):
        """Excel row indexes.

        StyleFrame row indexes (including the headers) according to the excel file format.
        Mostly used to set rows height.
        Excel indexes format starts from index 1.

        :rtype: tuple

        :meta private:
        """

        return tuple(range(1, len(self) + 2))

    def to_excel(self, excel_writer='output.xlsx', sheet_name='Sheet1',
                 allow_protection=False, right_to_left=False, columns_to_hide=None, row_to_add_filters=None,
                 columns_and_rows_to_freeze=None, best_fit=None, **kwargs):
        """Saves the dataframe to excel and applies the styles.

        .. note:: :meth:`to_excel` also accepts all arguments that :meth:`pandas.DataFrame.to_excel` accepts as kwargs.

        :param excel_writer: File path or existing ExcelWriter
        :type excel_writer: str or :class:`pandas.ExcelWriter` or :class:`pathlib.Path`
        :param str sheet_name: Name of sheet the StyleFrame will be exported to
        :param bool allow_protection: Allow to protect the cells that specified as protected. If used ``protection=True``
            in a Styler object this must be set to ``True``.
        :param bool right_to_left: Makes the sheet right-to-left.
        :param columns_to_hide: Columns names to hide.
        :type columns_to_hide: None or str or list or tuple or set
        :param row_to_add_filters: Add filters to the given row index, starts from 0 (which will add filters to header row).
        :type row_to_add_filters: None or int
        :param columns_and_rows_to_freeze: Column and row string to freeze.
            For example "C3" will freeze columns: A, B and rows: 1, 2.
        :type columns_and_rows_to_freeze: None or str

        .. versionadded:: 1.4

        :param best_fit: single column, list, set or tuple of columns names to attempt to best fit the width for.

            .. note:: ``best_fit`` will attempt to calculate the correct column-width based on the longest value in each provided
                      column. However this isn't guaranteed to work for all fonts (works best with monospaced fonts). The formula
                      used to calculate a column's width is equivalent to

                      ::

                        (len(longest_value_in_column) + A_FACTOR) * P_FACTOR

                      The default values for ``A_FACTOR`` and ``P_FACTOR`` are 13 and 1.3 respectively, and can be modified before
                      calling ``StyleFrame.to_excel`` by directly modifying ``StyleFrame.A_FACTOR`` and ``StyleFrame.P_FACTOR``

        :type best_fit: None or str or list or tuple or set
        :return: self
        :rtype: :class:`StyleFrame`

        """

        # dealing with needed pandas.to_excel defaults
        header = kwargs.pop('header', True)
        index = kwargs.pop('index', False)
        startcol = kwargs.pop('startcol', 0)
        startrow = kwargs.pop('startrow', 0)
        na_rep = kwargs.pop('na_rep', '')

        def get_values(x):
            if isinstance(x, Container):
                return x.value
            else:
                try:
                    if np.isnan(x):
                        return na_rep
                    else:
                        return x
                except TypeError:
                    return x

        def within_sheet_boundaries(row=1, column='A'):
            return (1 <= int(row) <= sheet.max_row
                        and
                    1 <= cell.column_index_from_string(column) <= sheet.max_column)

        def get_range_of_cells(row_index=None, columns=None):
            if columns is None:
                start_letter = self._get_column_as_letter(sheet, self.data_df.columns[0], startcol)
                end_letter = self._get_column_as_letter(sheet, self.data_df.columns[-1], startcol)
            else:
                start_letter = self._get_column_as_letter(sheet, columns[0], startcol)
                end_letter = self._get_column_as_letter(sheet, columns[-1], startcol)
            if row_index is None:  # returns cells range for the entire dataframe
                start_index = startrow + 1
                end_index = start_index + len(self)
            else:
                start_index = startrow + row_index + 1
                end_index = start_index
            return '{start_letter}{start_index}:{end_letter}{end_index}'.format(start_letter=start_letter,
                                                                                start_index=start_index,
                                                                                end_letter=end_letter,
                                                                                end_index=end_index)

        if len(self.data_df) > 0:
            export_df = self.data_df.applymap(get_values)

        else:
            export_df = deepcopy(self.data_df)

        export_df.columns = [col.value for col in export_df.columns]
        # noinspection PyTypeChecker
        export_df.index = [row_index.value for row_index in export_df.index]
        export_df.index.name = self.data_df.index.name

        if isinstance(excel_writer, (str, pathlib.Path)):
            excel_writer = self.ExcelWriter(excel_writer)

        export_df.to_excel(excel_writer, sheet_name=sheet_name, engine='openpyxl', header=header,
                           index=index, startcol=startcol, startrow=startrow, na_rep=na_rep, **kwargs)

        sheet = excel_writer.sheets[sheet_name]

        sheet.sheet_view.rightToLeft = right_to_left

        self.data_df.fillna(Container('NaN'), inplace=True)

        if index:
            if self.data_df.index.name:
                index_name_cell = sheet.cell(row=startrow + 1, column=startcol + 1)
                index_name_cell.style = self._index_header_style.to_openpyxl_style()
            for row_index, index in enumerate(self.data_df.index):
                try:
                    date_time_types_to_formats = {pd_timestamp: index.style.date_time_format,
                                                  dt.datetime: index.style.date_time_format,
                                                  dt.date: index.style.date_format,
                                                  dt.time: index.style.time_format}
                    index.style.number_format = date_time_types_to_formats.get(type(index.value),
                                                                               index.style.number_format)
                    style_to_apply = index.style.to_openpyxl_style()
                except AttributeError:
                    style_to_apply = index.style
                current_cell = sheet.cell(row=startrow + row_index + 2, column=startcol + 1)
                current_cell.style = style_to_apply
                if isinstance(index.style, Styler):
                    current_cell.comment = index.style.generate_comment()
                else:
                    if hasattr(index.style, 'comment'):
                        index.style.comment.parent = None
                        current_cell.comment = index.style.comment

            startcol += 1

        if header and not self._has_custom_headers_style:
            self.apply_headers_style(Styler.default_header_style())

        # Iterating over the dataframe's elements and applying their styles
        # openpyxl's rows and cols start from 1,1 while the dataframe is 0,0
        for col_index, column in enumerate(self.data_df.columns):
            try:
                date_time_types_to_formats = {pd_timestamp: column.style.date_time_format,
                                              dt.datetime: column.style.date_time_format,
                                              dt.date: column.style.date_format,
                                              dt.time: column.style.time_format}

                column.style.number_format = date_time_types_to_formats.get(type(column.value),
                                                                            column.style.number_format)
                style_to_apply = column.style.to_openpyxl_style()
            except AttributeError:
                style_to_apply = Styler.from_openpyxl_style(column.style, [],
                                                            openpyxl_comment=column.style.comment).to_openpyxl_style()
            column_header_cell = sheet.cell(row=startrow + 1, column=col_index + startcol + 1)
            column_header_cell.style = style_to_apply
            if isinstance(column.style, Styler):
                column_header_cell.comment = column.style.generate_comment()
            else:
                if hasattr(column.style, 'comment') and column.style.comment is not None:
                    column_header_cell.comment = column.style.comment
            for row_index, index in enumerate(self.data_df.index):
                current_cell = sheet.cell(row=row_index + startrow + (2 if header else 1), column=col_index + startcol + 1)
                data_df_style = self.data_df.at[index, column].style
                try:
                    if '=HYPERLINK' in str(current_cell.value):
                        data_df_style.font_color = utils.colors.blue
                        data_df_style.underline = utils.underline.single
                    else:
                        if best_fit and column.value in best_fit:
                            data_df_style.wrap_text = False
                            data_df_style.shrink_to_fit = False
                    try:
                        date_time_types_to_formats = {pd_timestamp: data_df_style.date_time_format,
                                                      dt.datetime: data_df_style.date_time_format,
                                                      dt.date: data_df_style.date_format,
                                                      dt.time: data_df_style.time_format}

                        data_df_style.number_format = date_time_types_to_formats.get(type(self.data_df.at[index,column].value),
                                                                                     data_df_style.number_format)
                        style_to_apply = data_df_style.to_openpyxl_style()
                    except AttributeError:
                        style_to_apply = Styler.from_openpyxl_style(data_df_style, [],
                                                                    openpyxl_comment=data_df_style.comment).to_openpyxl_style()
                    current_cell.style = style_to_apply
                    if isinstance(data_df_style, Styler):
                        current_cell.comment = data_df_style.generate_comment()
                    else:
                        if hasattr(data_df_style, 'comment') and data_df_style.comment is not None:
                            current_cell.comment = data_df_style.comment
                except AttributeError:  # if the element in the dataframe is not Container creating a default style
                    current_cell.style = Styler().to_openpyxl_style()

        if best_fit:
            if not isinstance(best_fit, (list, set, tuple)):
                best_fit = [best_fit]
            self.set_column_width_dict({column: (max(self.data_df[column].astype(str).str.len()) + self.A_FACTOR) * self.P_FACTOR
                                        for column in best_fit})

        for column in self._columns_width:
            column_letter = self._get_column_as_letter(sheet, column, startcol)
            sheet.column_dimensions[column_letter].width = self._columns_width[column]

        for row in self._rows_height:
            if within_sheet_boundaries(row=(row + startrow)):
                sheet.row_dimensions[startrow + row].height = self._rows_height[row]
            else:
                raise IndexError('row: {} is out of range'.format(row))

        if row_to_add_filters is not None:
            try:
                row_to_add_filters = int(row_to_add_filters)
                if not within_sheet_boundaries(row=(row_to_add_filters + startrow + 1)):
                    raise IndexError('row: {} is out of rows range'.format(row_to_add_filters))
                sheet.auto_filter.ref = get_range_of_cells(row_index=row_to_add_filters)
            except (TypeError, ValueError):
                raise TypeError("row must be an index and not {}".format(type(row_to_add_filters)))

        if columns_and_rows_to_freeze is not None:
            if not isinstance(columns_and_rows_to_freeze, str) or len(columns_and_rows_to_freeze) < 2:
                raise TypeError("columns_and_rows_to_freeze must be a str for example: 'C3'")
            if not within_sheet_boundaries(column=columns_and_rows_to_freeze[0]):
                raise IndexError("column: %s is out of columns range." % columns_and_rows_to_freeze[0])
            if not within_sheet_boundaries(row=columns_and_rows_to_freeze[1]):
                raise IndexError("row: %s is out of rows range." % columns_and_rows_to_freeze[1])
            sheet.freeze_panes = sheet[columns_and_rows_to_freeze]

        if allow_protection:
            sheet.protection.autoFilter = False
            sheet.protection.enable()

        # Iterating over the columns_to_hide and check if the format is columns name, column index as number or letter
        if columns_to_hide:
            if not isinstance(columns_to_hide, (list, set, tuple)):
                columns_to_hide = [columns_to_hide]

            for column in columns_to_hide:
                column_letter = self._get_column_as_letter(sheet, column, startcol)
                sheet.column_dimensions[column_letter].hidden = True

        for cond_formatting in self._cond_formatting:
            sheet.conditional_formatting.add(get_range_of_cells(columns=cond_formatting.columns),
                                             cond_formatting.rule)

        return excel_writer

    def apply_style_by_indexes(self, indexes_to_style, styler_obj, cols_to_style=None, height=None,
                               complement_style=None, complement_height=None, overwrite_default_style=True):
        """
        Applies a certain style to the provided indexes in the dataframe in the provided columns

        :param indexes_to_style: Indexes to which the provided style will be applied.
            Usually passed as pandas selecting syntax. For example,

            ::

                sf[sf['some_col'] = 20]

        :type indexes_to_style: list or tuple or int or Container
        :param styler_obj: `Styler` object that contains the style that will be applied to indexes in `indexes_to_style`
        :type styler_obj: :class:`.Styler`
        :param cols_to_style: The column names to apply the provided style to. If ``None`` all columns will be styled.
        :type cols_to_style: None or str or list[str] or tuple[str] or set[str]
        :param height: If provided, set height for rows whose indexes are in `indexes_to_style`.
        :type height: None or int or float

        .. versionadded:: 1.5

        :param complement_style: `Styler` object that contains the style which will be applied to indexes not in `indexes_to_style`
        :type complement_style: None or :class:`.Styler`
        :param complement_height: Height for rows whose indexes are not in `indexes_to_style`. If not provided then
                `height` will be used (if provided).
        :type complement_height: None or int or float

        .. versionadded:: 1.6

        :param bool overwrite_default_style: If ``True``, the default style (the style used when initializing StyleFrame)
                will be overwritten. If ``False`` then the default style and the provided style wil be combined using
                :meth:`.Styler.combine` method.

        :return: self
        :rtype: :class:`StyleFrame`
        """

        if not isinstance(styler_obj, Styler):
            raise TypeError('styler_obj must be {}, got {} instead.'.format(Styler.__name__, type(styler_obj).__name__))

        if isinstance(indexes_to_style, (list, tuple, int)):
            indexes_to_style = self.index[indexes_to_style]

        elif isinstance(indexes_to_style, Container):
            indexes_to_style = pd.Index([indexes_to_style])

        if cols_to_style is not None and not isinstance(cols_to_style, (list, tuple, set)):
            cols_to_style = [cols_to_style]
        elif cols_to_style is None:
            cols_to_style = list(self.data_df.columns)

        if overwrite_default_style:
            style_to_apply = deepcopy(styler_obj)
        else:
            style_to_apply = Styler.combine(self._default_style, styler_obj)

        for index in indexes_to_style:
            index.style = style_to_apply
            for col in cols_to_style:
                self.iloc[self.index.get_loc(index), self.columns.get_loc(col)].style = style_to_apply

        if height:
            # Add offset 2 since rows do not include the headers and they starts from 1 (not 0).
            rows_indexes_for_height_change = [self.index.get_loc(idx) + 2 for idx in indexes_to_style]
            self.set_row_height(rows=rows_indexes_for_height_change, height=height)

        if complement_style:
            self.apply_style_by_indexes(self.index.difference(indexes_to_style), complement_style, cols_to_style,
                                        complement_height if complement_height else height)

        return self

    def apply_column_style(self, cols_to_style, styler_obj, style_header=False, use_default_formats=True, width=None,
                           overwrite_default_style=True):
        """Apply style to a whole column

        :param cols_to_style: The column names to style.
        :type cols_to_style: str or list or tuple or set
        :param styler_obj: A `Styler` object.
        :type styler_obj: :class:`.Styler`
        :param bool style_header: If ``True``, the column(s) header will also be styled.
        :param bool use_default_formats: If ``True``, the default formats for date and times will be used.
        :param width: If provided, the new width for the specified columns.
        :type width: None or int or float
        :param bool overwrite_default_style: (bool) If ``True``, the default style (the style used when initializing StyleFrame)
                will be overwritten. If ``False`` then the default style and the provided style wil be combined using
                :meth:`.Styler.combine` method.
        :return: self
        :rtype: :class:`StyleFrame`
        """

        if not isinstance(styler_obj, Styler):
            raise TypeError('styler_obj must be {}, got {} instead.'.format(Styler.__name__, type(styler_obj).__name__))

        if not isinstance(cols_to_style, (list, tuple, set, pd.Index)):
            cols_to_style = [cols_to_style]
        if not all(col in self.columns for col in cols_to_style):
            raise KeyError("one of the columns in {} wasn't found".format(cols_to_style))

        if overwrite_default_style:
            style_to_apply = styler_obj
        else:
            style_to_apply = Styler.combine(self._default_style, styler_obj)

        for col_name in cols_to_style:
            if style_header:
                self.columns[self.columns.get_loc(col_name)].style = style_to_apply
                self._has_custom_headers_style = True
            for index in self.index:
                if use_default_formats:
                    if isinstance(self.at[index, col_name].value, pd_timestamp):
                        style_to_apply.number_format = utils.number_formats.date_time
                    elif isinstance(self.at[index, col_name].value, dt.date):
                        style_to_apply.number_format = utils.number_formats.date
                    elif isinstance(self.at[index, col_name].value, dt.time):
                        style_to_apply.number_format = utils.number_formats.time_24_hours

                self.at[index, col_name].style = style_to_apply

        if width:
            self.set_column_width(columns=cols_to_style, width=width)

        return self

    def apply_headers_style(self, styler_obj, style_index_header=True, cols_to_style=None):
        """Apply style to the headers only

        :param styler_obj: The style to apply
        :type styler_obj: :class:`.Styler`

        .. versionadded:: 1.6.1

        :param bool style_index_header: If True then the style will also be applied to the header of the index column

        .. versionadded:: 2.0.5

        :param cols_to_style: the columns to apply the style to, if not provided all the columns will be styled
        :type cols_to_style: None or str or list[str] or tuple[str] or set[str]

        :return: self
        :rtype: :class:`StyleFrame`
        """

        if not isinstance(styler_obj, Styler):
            raise TypeError('styler_obj must be {}, got {} instead.'.format(Styler.__name__, type(styler_obj).__name__))

        if cols_to_style is None:
            cols_to_style = self.data_df.columns
        if not isinstance(cols_to_style, (list, tuple, set, pd.Index)):
            cols_to_style = [cols_to_style]
        if not all(col in self.columns for col in cols_to_style):
            raise KeyError("one of the columns in {} wasn't found".format(cols_to_style))

        if style_index_header:
            self._index_header_style = styler_obj

        for column in cols_to_style:
            self.columns[self.columns.get_loc(column)].style = styler_obj
        self._has_custom_headers_style = True
        return self

    def set_column_width(self, columns, width):
        """Set the width of the given columns

        :param columns: Column name(s) or index(es).
        :type columns: str or list[str] or tuple[str] or int or list[int] or tuple[int]
        :param width: The new width for the specified columns.
        :type width: int or float
        :return: self
        :rtype: :class:`StyleFrame`
        """

        if not isinstance(columns, (set, list, tuple, pd.Index)):
            columns = [columns]
        try:
            width = float(width)
        except ValueError:
            raise TypeError('columns width must be numeric value')

        if width <= 0:
            raise ValueError('columns width must be positive')

        for column in columns:
            if not isinstance(column, (int, str, Container)):
                raise TypeError("column must be an index, column letter or column name")
            self._columns_width[column] = width

        return self

    def set_column_width_dict(self, col_width_dict):
        """
        :param col_width_dict: A dictionary from column names to width.
        :type col_width_dict: dict[str, int or float]
        :return: self
        :rtype: :class:`StyleFrame`
        """

        if not isinstance(col_width_dict, dict):
            raise TypeError("'col_width_dict' must be a dictionary")
        for cols, width in col_width_dict.items():
            self.set_column_width(cols, width)

        return self

    def set_row_height(self, rows, height):
        """ Set the height of the given rows

        :param rows: Row(s) index.
        :type rows: int or list[int] or tuple[int] or set[int]
        :param height: The new height for the specified indexes.
        :type height: int or float
        :return: self
        :rtype: :class:`StyleFrame`
        """

        if not isinstance(rows, (set, list, tuple, pd.Index)):
            rows = [rows]
        try:
            height = float(height)
        except ValueError:
            raise TypeError('rows height must be numeric value')

        if height <= 0:
            raise ValueError('rows height must be positive')
        for row in rows:
            try:
                row = int(row)
            except TypeError:
                raise TypeError("row must be an index")

            self._rows_height[row] = height

        return self

    def set_row_height_dict(self, row_height_dict):
        """
        :param row_height_dict: A dictionary from row indexes to height.
        :type row_height_dict: dict[int, int or float]
        :return: self
        :rtype: :class:`StyleFrame`
        """

        if not isinstance(row_height_dict, dict):
            raise TypeError("'row_height_dict' must be a dictionary")
        for rows, height in row_height_dict.items():
            self.set_row_height(rows, height)
        return self

    def rename(self, columns=None, inplace=False):
        """Renames the underlying dataframe's columns

        :param dict columns: A dictionary from old columns names to new columns names.
        :param bool inplace: If ``False``, a new StyleFrame object will be returned. If ``True``, renames the columns inplace.
        :return: self if inplace is ``True``, new StyleFrame object is ``False``
        :rtype: :class:`StyleFrame`
        """

        if not isinstance(columns, dict):
            raise TypeError("'columns' must be a dictionary")

        sf = self if inplace else StyleFrame(self)

        new_columns = [col if col not in columns else Container(columns[col], col.style)
                       for col in sf.data_df.columns]

        sf._known_attrs['columns'] = sf.data_df.columns = new_columns

        sf._columns_width.update({new_col_name: sf._columns_width.pop(old_col_name)
                                  for old_col_name, new_col_name in columns.items()
                                  if old_col_name in sf._columns_width})
        return sf

    def style_alternate_rows(self, styles, **kwargs):
        """
        .. versionadded:: 1.2

        Applies the provided styles to rows in an alternating manner.

        .. note:: :meth:`style_alternate_rows` also accepts all arguments that :meth:`apply_style_by_indexes` accepts as kwargs.

        :param styles: List, tuple or set of :class:`.Styler` objects to be applied to rows in an alternating manner
        :type styles: list[:class:`.Styler`] or tuple[:class:`.Styler`] or set[:class:`.Styler`]
        :return: self
        :rtype: :class:`StyleFrame`

        """

        num_of_styles = len(styles)
        split_indexes = (self.index[i::num_of_styles] for i in range(num_of_styles))
        for i, indexes in enumerate(split_indexes):
            self.apply_style_by_indexes(indexes, styles[i], **kwargs)
        return self

    def add_color_scale_conditional_formatting(self, start_type, start_value, start_color, end_type, end_value, end_color,
                                               mid_type=None, mid_value=None, mid_color=None, columns_range=None):
        """
        :param start_type: The type for the minimum bound
        :type start_type: str: one of :class:`.utils.conditional_formatting_types` or any other type Excel supports
        :param start_value: The threshold for the minimum bound
        :param start_color: The color for the minimum bound
        :type start_color: str: one of :class:`.utils.colors`, hex string or color name ie `'yellow'` Excel supports
        :param end_type: The type for the maximum bound
        :type end_type: str: one of :class:`.utils.conditional_formatting_types` or any other type Excel supports
        :param end_value: The threshold for the maximum bound
        :param end_color: The color for the maximum bound
        :type end_color: str: one of :class:`.utils.colors`, hex string or color name ie `'yellow'` Excel supports
        :param mid_type: The type for the middle bound
        :type mid_type: None or str: one of :class:`.utils.conditional_formatting_types` or any other type Excel supports
        :param mid_value: The threshold for the middle bound
        :param mid_color: The color for the middle bound
        :type mid_color: None or str: one of :class:`.utils.colors`, hex string or color name ie `'yellow'` Excel supports
        :param columns_range: A two-elements list or tuple of columns to which the conditional formatting will be added
                to.
                If not provided at all the conditional formatting will be added to all columns.
                If a single element is provided then the conditional formatting will be added to the provided column.
                If two elements are provided then the conditional formatting will start in the first column and end in the second.
                The provided columns can be a column name, letter or index.
        :type columns_range: None or list[str or int] or tuple[str or int])
        :return: self
        :rtype: :class:`StyleFrame`
        """

        if columns_range is None:
            columns_range = (self.data_df.columns[0], self.data_df.columns[-1])

        if not isinstance(columns_range, (list, tuple)) or len(columns_range) not in (1, 2):
            raise TypeError("'columns_range' should be a list or a tuple with 1 or 2 elements")

        self._cond_formatting.append(ColorScaleConditionalFormatRule(start_type=start_type, start_value=start_value,
                                                                     start_color=start_color,
                                                                     mid_type=mid_type, mid_value=mid_value,
                                                                     mid_color=mid_color,
                                                                     end_type=end_type, end_value=end_value,
                                                                     end_color=end_color,
                                                                     columns_range=columns_range))

        return self
