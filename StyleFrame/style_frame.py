# coding:utf-8

import datetime as dt
import numpy as np
import pandas as pd
import sys

from .deprecations import deprecated_kwargs
from . import utils
from copy import deepcopy
from collections import Iterable
from openpyxl import load_workbook
from openpyxl.cell.cell import get_column_letter
from openpyxl.xml.functions import fromstring, QName
from openpyxl.utils import cell

PY2 = sys.version_info[0] == 2

# Python 2
if PY2:
    # noinspection PyUnresolvedReferences
    from container import Container
    # noinspection PyUnresolvedReferences
    from series import Series
    # noinspection PyUnresolvedReferences
    from styler import Styler, ColorScaleConditionalFormatRule

# Python 3
else:
    from StyleFrame.container import Container
    from StyleFrame.styler import Styler, ColorScaleConditionalFormatRule
    from StyleFrame.series import Series

try:
    pd_timestamp = pd.Timestamp
except AttributeError:
    pd_timestamp = pd.tslib.Timestamp

str_type = basestring if PY2 else str
unicode_type = unicode if PY2 else str


class StyleFrame(object):
    """
    A wrapper class that wraps pandas DataFrame.
    Stores container objects that have values and Styles that will be applied to excel
    """
    P_FACTOR = 1.3
    A_FACTOR = 13

    def __init__(self, obj, styler_obj=None):
        from_another_styleframe = False
        from_pandas_dataframe = False
        if styler_obj and not isinstance(styler_obj, Styler):
            raise TypeError('styler_obj must be {}, got {} instead.'.format(Styler.__name__, type(styler_obj).__name__))
        if isinstance(obj, pd.DataFrame):
            from_pandas_dataframe = True
            if obj.empty:
                self.data_df = deepcopy(obj)
            else:
                self.data_df = obj.applymap(lambda x: Container(x, deepcopy(styler_obj)) if not isinstance(x, Container) else x)
        elif isinstance(obj, pd.Series):
            self.data_df = obj.apply(lambda x: Container(x, deepcopy(styler_obj)) if not isinstance(x, Container) else x)
        elif isinstance(obj, (dict, list)):
            self.data_df = pd.DataFrame(obj).applymap(lambda x: Container(x, deepcopy(styler_obj)) if not isinstance(x, Container) else x)
        elif isinstance(obj, StyleFrame):
            self.data_df = deepcopy(obj.data_df)
            from_another_styleframe = True
        else:
            raise TypeError("{} __init__ doesn't support {}".format(type(self).__name__, type(obj).__name__))
        self.data_df.columns = [Container(col, deepcopy(styler_obj)) if not isinstance(col, Container) else deepcopy(col)
                                for col in self.data_df.columns]
        self.data_df.index = [Container(index, deepcopy(styler_obj)) if not isinstance(index, Container) else deepcopy(index)
                              for index in self.data_df.index]

        if from_pandas_dataframe:
            self.data_df.index.name = obj.index.name

        self._columns_width = obj._columns_width if from_another_styleframe else {}
        self._rows_height = obj._rows_height if from_another_styleframe else {}
        self._has_custom_headers_style = obj._has_custom_headers_style if from_another_styleframe else False
        self._cond_formatting = []
        self._default_style = styler_obj or Styler()
        self._index_header_style = obj._index_header_style if from_another_styleframe else self._default_style

        self._known_attrs = {'at': self.data_df.at,
                             'loc': self.data_df.loc,
                             'iloc': self.data_df.iloc,
                             'applymap': self.data_df.applymap,
                             'groupby': self.data_df.groupby,
                             'index': self.data_df.index,
                             'columns': self.data_df.columns,
                             'fillna': self.data_df.fillna}

    def __str__(self):
        return str(self.data_df)

    def __unicode__(self):
        return unicode_type(self.data_df)

    def __len__(self):
        return len(self.data_df)

    def __getitem__(self, item):
        if isinstance(item, pd.Series):
            return self.data_df.__getitem__(item).index
        if isinstance(item, list):
            return StyleFrame(self.data_df.__getitem__(item))
        return Series(self.data_df.__getitem__(item))

    def __setitem__(self, key, value):
        if isinstance(value, (Iterable, pd.Series)):
            self.data_df.__setitem__(Container(key), list(map(Container, value)))
        else:
            self.data_df.__setitem__(Container(key), Container(value))

    def __delitem__(self, item):
        return self.data_df.__delitem__(item)

    def __getattr__(self, attr):
        if attr in self.data_df.columns:
            return self.data_df[attr]
        try:
            return self._known_attrs[attr]
        except KeyError:
            raise AttributeError("'{}' object has no attribute '{}'".format(type(self).__name__, attr))

    def _get_column_as_letter(self, sheet, column_to_convert, startcol=0):
        if not isinstance(column_to_convert, (int, str_type, unicode_type, Container)):
            raise TypeError("column must be an index, column letter or column name")
        column_as_letter = None
        if column_to_convert in self.data_df.columns:  # column name
            column_index = self.data_df.columns.get_loc(
                column_to_convert) + startcol + 1  # worksheet columns index start from 1
            column_as_letter = cell.get_column_letter(column_index)

        # column index
        elif isinstance(column_to_convert, int) and column_to_convert >= 1:
            column_as_letter = cell.get_column_letter(startcol + column_to_convert)

        # assuming we got column letter
        elif isinstance(column_to_convert, (str_type, unicode_type)) < get_column_letter(sheet.max_column):
            column_as_letter = column_to_convert

        if column_as_letter is None or cell.column_index_from_string(column_as_letter) > sheet.max_column:
            raise IndexError("column: %s is out of columns range." % column_to_convert)

        return column_as_letter

    @classmethod
    @deprecated_kwargs(('sheetname',))
    def read_excel(cls, path, sheet_name=0, read_style=False, use_openpyxl_styles=False,
                   read_comments=False, **kwargs):
        """Creates a StyleFrame object from an existing Excel.

        :param str path: The path to the Excel file to read.
        :param str|int sheet_name: The sheet name to read. If an integer is provided then it be used as a zero-based
            sheet index
        :param bool read_style: If True the sheet's style will be loaded to the returned StyleFrame object.
        :param bool use_openpyxl_styles: If True (and read_style is also True) then the styles in the returned
            StyleFrame object will be Openpyxl's style objects. If False, the styles will be StyleFrame.Styler objects.
            Defaults to True for backward compatibility.
        :param bool read_comments: If True cells' comments will be loaded to the returned StyleFrame object. Note
            that reading comments without reading styles is currently not supported.
        :param kwargs: Any keyword argument pandas' `read_excel` supports.
        :rtype: StyleFrame
        """

        def _get_scheme_colors_from_excel(wb):
            xlmns = 'http://schemas.openxmlformats.org/drawingml/2006/main'
            if wb.loaded_theme is None:
                return []
            root = fromstring(wb.loaded_theme)
            theme_element = root.find(QName(xlmns, 'themeElements').text)
            color_schemes = theme_element.findall(QName(xlmns, 'clrScheme').text)
            colors = []
            for colorScheme in color_schemes:
                for tag in ['lt1', 'dk1', 'lt2', 'dk2', 'accent1', 'accent2', 'accent3', 'accent4', 'accent5', 'accent6']:
                    accent = list(colorScheme.find(QName(xlmns, tag).text))[0]
                    if 'window' in accent.attrib['val']:
                        colors.append(accent.attrib['lastClr'])
                    else:
                        colors.append(accent.attrib['val'])
            return colors

        def _read_style():
            wb = load_workbook(path)
            if isinstance(sheet_name, (str_type, unicode_type)):
                sheet = wb[sheet_name]
            elif isinstance(sheet_name, int):
                sheet = wb.worksheets[sheet_name]
            else:
                raise TypeError("'sheet_name' must be a string or int, got {} instead".format(type(sheet_name)))
            theme_colors = _get_scheme_colors_from_excel(wb)
            for col_index, col_name in enumerate(sf.columns, start=1):
                column_cell = sheet.cell(row=1, column=col_index)
                if use_openpyxl_styles:
                    style_object = column_cell
                else:
                    style_object = Styler.from_openpyxl_style(column_cell, theme_colors,
                                                              read_comments and column_cell.comment)
                sf.columns[col_index - 1].style = style_object
                for row_index, sf_index in enumerate(sf.index, start=2):
                    current_cell = sheet.cell(row=row_index, column=col_index)
                    if use_openpyxl_styles:
                        style_object = current_cell
                    else:
                        style_object = Styler.from_openpyxl_style(current_cell, theme_colors,
                                                                  read_comments and current_cell.comment)
                    sf.at[sf_index, col_name].style = style_object
                    sf._rows_height[row_index] = sheet.row_dimensions[row_index].height

                sf._columns_width[col_name] = sheet.column_dimensions[sf._get_column_as_letter(sheet, col_name)].width

        if 'sheetname' in kwargs:
            sheet_name = kwargs.pop('sheetname')

        sf = cls(pd.read_excel(path, sheet_name, **kwargs))
        if read_style:
            _read_style()
            sf._has_custom_headers_style = True
        return sf

    # noinspection PyPep8Naming
    @classmethod
    def ExcelWriter(cls, path):
        return pd.ExcelWriter(path, engine='openpyxl')

    @property
    def row_indexes(self):
        """Excel row indexes.

        StyleFrame row indexes (including the headers) according to the excel file format.
        Mostly used to set rows height.
        Excel indexes format starts from index 1.

        :rtype: tuple
        """

        return tuple(range(1, len(self) + 2))

    def to_excel(self, excel_writer='output.xlsx', sheet_name='Sheet1',
                 allow_protection=False, right_to_left=False, columns_to_hide=None, row_to_add_filters=None,
                 columns_and_rows_to_freeze=None, best_fit=None, **kwargs):
        """Saves the dataframe to excel and applies the styles.

        :param str|pandas.ExcelWriter excel_writer: File path or existing ExcelWriter
        :param str sheet_name: Name of sheet the StyleFrame will be exported to
        :param bool right_to_left: sets the sheet to be right to left.
        :param None|str|list|tuple|set columns_to_hide: single column, list, set or tuple of columns to hide, may be column index (starts from 1)
                                column name or column letter.
        :param bool allow_protection: allow to protect the sheet and the cells that specified as protected.
        :param None|int row_to_add_filters: add filters to the given row, starts from zero (zero is to add filters to columns).
        :param None|str columns_and_rows_to_freeze: column and row string to freeze for example: C3 will freeze columns: A,B and rows: 1,2.
        :param None|str|list|tuple|set best_fit: single column, list, set or tuple of columns names to attempt to best fit the width
                                for.

        See Pandas.DataFrame.to_excel documentation about other arguments
        """

        # dealing with needed pandas.to_excel defaults
        header = kwargs.pop('header', True)
        index = kwargs.pop('index', False)
        startcol = kwargs.pop('startcol', 0)
        startrow = kwargs.pop('startrow', 0)
        na_rep = kwargs.pop('na_rep', '')

        def get_values(x):
            if isinstance(x, Container):
                return x.value
            else:
                try:
                    if np.isnan(x):
                        return na_rep
                    else:
                        return x
                except TypeError:
                    return x

        def within_sheet_boundaries(row=1, column='A'):
            return (1 <= int(row) <= sheet.max_row
                        and
                    1 <= cell.column_index_from_string(column) <= sheet.max_column)

        def get_range_of_cells(row_index=None, columns=None):
            if columns is None:
                start_letter = self._get_column_as_letter(sheet, self.data_df.columns[0], startcol)
                end_letter = self._get_column_as_letter(sheet, self.data_df.columns[-1], startcol)
            else:
                start_letter = self._get_column_as_letter(sheet, columns[0], startcol)
                end_letter = self._get_column_as_letter(sheet, columns[-1], startcol)
            if row_index is None:  # returns cells range for the entire dataframe
                start_index = startrow + 1
                end_index = start_index + len(self)
            else:
                start_index = startrow + row_index + 1
                end_index = start_index
            return '{start_letter}{start_index}:{end_letter}{end_index}'.format(start_letter=start_letter,
                                                                                start_index=start_index,
                                                                                end_letter=end_letter,
                                                                                end_index=end_index)

        if len(self.data_df) > 0:
            export_df = self.data_df.applymap(get_values)

        else:
            export_df = deepcopy(self.data_df)

        export_df.columns = [col.value for col in export_df.columns]
        # noinspection PyTypeChecker
        export_df.index = [row_index.value for row_index in export_df.index]
        export_df.index.name = self.data_df.index.name

        if isinstance(excel_writer, (str_type, unicode_type)):
            excel_writer = self.ExcelWriter(excel_writer)

        export_df.to_excel(excel_writer, sheet_name=sheet_name, engine='openpyxl', header=header,
                           index=index, startcol=startcol, startrow=startrow, na_rep=na_rep, **kwargs)

        sheet = excel_writer.sheets[sheet_name]

        sheet.sheet_view.rightToLeft = right_to_left

        self.data_df.fillna(Container('NaN'), inplace=True)

        if index:
            if self.data_df.index.name:
                index_name_cell = sheet.cell(row=startrow + 1, column=startcol + 1)
                index_name_cell.style = self._index_header_style.to_openpyxl_style()
            for row_index, index in enumerate(self.data_df.index):
                try:
                    style_to_apply = index.style.to_openpyxl_style()
                except AttributeError:
                    style_to_apply = index.style
                current_cell = sheet.cell(row=startrow + row_index + 2, column=startcol + 1)
                current_cell.style = style_to_apply
                if isinstance(index.style, Styler):
                    current_cell.comment = index.style.generate_comment()
                else:
                    if hasattr(index.style, 'comment'):
                        index.style.comment.parent = None
                        current_cell.comment = index.style.comment

            startcol += 1

        if header and not self._has_custom_headers_style:
            self.apply_headers_style(Styler.default_header_style())

        # Iterating over the dataframe's elements and applying their styles
        # openpyxl's rows and cols start from 1,1 while the dataframe is 0,0
        for col_index, column in enumerate(self.data_df.columns):
            try:
                style_to_apply = column.style.to_openpyxl_style()
            except AttributeError:
                style_to_apply = column.style
            column_header_cell = sheet.cell(row=startrow + 1, column=col_index + startcol + 1)
            column_header_cell.style = style_to_apply
            if isinstance(column.style, Styler):
                column_header_cell.comment = column.style.generate_comment()
            else:
                if hasattr(column.style, 'comment'):
                    column.style.comment.parent = None
                    column_header_cell.comment = column.style.comment
            for row_index, index in enumerate(self.data_df.index):
                current_cell = sheet.cell(row=row_index + startrow + 2, column=col_index + startcol + 1)
                data_df_style = self.data_df.at[index, column].style
                try:
                    if '=HYPERLINK' in unicode_type(current_cell.value):
                        data_df_style.font_color = utils.colors.blue
                        data_df_style.underline = utils.underline.single
                    else:
                        if best_fit and column.value in best_fit:
                            data_df_style.wrap_text = False
                            data_df_style.shrink_to_fit = False
                    try:
                        style_to_apply = data_df_style.to_openpyxl_style()
                    except AttributeError:
                        style_to_apply = data_df_style
                    current_cell.style = style_to_apply
                    if isinstance(data_df_style, Styler):
                        current_cell.comment = data_df_style.generate_comment()
                    else:
                        if hasattr(data_df_style, 'comment'):
                            data_df_style.comment.parent = None
                            current_cell.comment = data_df_style.comment
                except AttributeError:  # if the element in the dataframe is not Container creating a default style
                    current_cell.style = Styler().to_openpyxl_style()

        if best_fit:
            if not isinstance(best_fit, (list, set, tuple)):
                best_fit = [best_fit]
            self.set_column_width_dict({column: (max(self.data_df[column].astype(str).str.len()) + self.A_FACTOR) * self.P_FACTOR
                                        for column in best_fit})

        for column in self._columns_width:
            column_letter = self._get_column_as_letter(sheet, column, startcol)
            sheet.column_dimensions[column_letter].width = self._columns_width[column]

        for row in self._rows_height:
            if within_sheet_boundaries(row=(row + startrow)):
                sheet.row_dimensions[startrow + row].height = self._rows_height[row]
            else:
                raise IndexError('row: {} is out of range'.format(row))

        if row_to_add_filters is not None:
            try:
                row_to_add_filters = int(row_to_add_filters)
                if not within_sheet_boundaries(row=(row_to_add_filters + startrow + 1)):
                    raise IndexError('row: {} is out of rows range'.format(row_to_add_filters))
                sheet.auto_filter.ref = get_range_of_cells(row_index=row_to_add_filters)
            except (TypeError, ValueError):
                raise TypeError("row must be an index and not {}".format(type(row_to_add_filters)))

        if columns_and_rows_to_freeze is not None:
            if not isinstance(columns_and_rows_to_freeze, (str_type, unicode_type)) or len(columns_and_rows_to_freeze) < 2:
                raise TypeError("columns_and_rows_to_freeze must be a str for example: 'C3'")
            if not within_sheet_boundaries(column=columns_and_rows_to_freeze[0]):
                raise IndexError("column: %s is out of columns range." % columns_and_rows_to_freeze[0])
            if not within_sheet_boundaries(row=columns_and_rows_to_freeze[1]):
                raise IndexError("row: %s is out of rows range." % columns_and_rows_to_freeze[1])
            sheet.freeze_panes = sheet[columns_and_rows_to_freeze]

        if allow_protection:
            sheet.protection.autoFilter = False
            sheet.protection.enable()

        # Iterating over the columns_to_hide and check if the format is columns name, column index as number or letter
        if columns_to_hide:
            if not isinstance(columns_to_hide, (list, set, tuple)):
                columns_to_hide = [columns_to_hide]

            for column in columns_to_hide:
                column_letter = self._get_column_as_letter(sheet, column, startcol)
                sheet.column_dimensions[column_letter].hidden = True

        for cond_formatting in self._cond_formatting:
            sheet.conditional_formatting.add(get_range_of_cells(columns=cond_formatting.columns),
                                             cond_formatting.rule)

        return excel_writer

    def apply_style_by_indexes(self, indexes_to_style, styler_obj, cols_to_style=None, height=None,
                               complement_style=None, complement_height=None, overwrite_default_style=True):
        """Applies a certain style to the provided indexes in the dataframe in the provided columns

        :param list|tuple|int|Container indexes_to_style: indexes to which the provided style will be applied
        :param Styler styler_obj: the styler object that contains the style which will be applied to indexes in indexes_to_style
        :param None|str|list|tuple|set cols_to_style: the columns to apply the style to, if not provided all the columns will be styled
        :param None|int|float height: height for rows whose indexes are in indexes_to_style
        :param None|Styler complement_style: the styler object that contains the style which will be applied to indexes not in indexes_to_style
        :param None|int|float complement_height: height for rows whose indexes are not in indexes_to_style. If not provided then
            height will be used (if provided).
        :param bool overwrite_default_style: If True, the default style (the style used when initializing StyleFrame)
            will be overwritten. If False then the default style and the provided style wil be combined using
            Styler.combine method.
        :return: self
        :rtype: StyleFrame
        """

        if not isinstance(styler_obj, Styler):
            raise TypeError('styler_obj must be {}, got {} instead.'.format(Styler.__name__, type(styler_obj).__name__))

        if isinstance(indexes_to_style, (list, tuple, int)):
            indexes_to_style = self.index[indexes_to_style]

        elif isinstance(indexes_to_style, Container):
            indexes_to_style = pd.Index([indexes_to_style])

        default_number_formats = {pd_timestamp: utils.number_formats.default_date_time_format,
                                  dt.date: utils.number_formats.default_date_format,
                                  dt.time: utils.number_formats.default_time_format}

        orig_number_format = styler_obj.number_format

        if cols_to_style is not None and not isinstance(cols_to_style, (list, tuple, set)):
            cols_to_style = [cols_to_style]
        elif cols_to_style is None:
            cols_to_style = list(self.data_df.columns)

        if overwrite_default_style:
            style_to_apply = deepcopy(styler_obj)
        else:
            style_to_apply = Styler.combine(self._default_style, styler_obj)

        for index in indexes_to_style:
            if orig_number_format == utils.number_formats.general:
                style_to_apply.number_format = default_number_formats.get(type(index.value),
                                                                          utils.number_formats.general)
            index.style = style_to_apply

            for col in cols_to_style:
                cell = self.iloc[self.index.get_loc(index), self.columns.get_loc(col)]
                if orig_number_format == utils.number_formats.general:
                    style_to_apply.number_format = default_number_formats.get(type(cell.value),
                                                                              utils.number_formats.general)

                cell.style = style_to_apply

        if height:
            # Add offset 2 since rows do not include the headers and they starts from 1 (not 0).
            rows_indexes_for_height_change = [self.index.get_loc(idx) + 2 for idx in indexes_to_style]
            self.set_row_height(rows=rows_indexes_for_height_change, height=height)

        if complement_style:
            self.apply_style_by_indexes(self.index.difference(indexes_to_style), complement_style, cols_to_style,
                                        complement_height if complement_height else height)

        return self

    def apply_column_style(self, cols_to_style, styler_obj, style_header=False, use_default_formats=True, width=None,
                           overwrite_default_style=True):
        """apply style to a whole column

        :param str|list|tuple|set cols_to_style: the columns to apply the style to
        :param Styler styler_obj: the styler object that contains the style to be applied
        :param bool style_header: if True, style the headers as well
        :param bool use_default_formats: if True, use predefined styles for dates and times
        :param None|int|float width: non-default width for the given columns
        :param bool overwrite_default_style: If True, the default style (the style used when initializing StyleFrame)
            will be overwritten. If False then the default style and the provided style wil be combined using
            Styler.combine method.
        :return: self
        :rtype: StyleFrame
        """

        if not isinstance(styler_obj, Styler):
            raise TypeError('styler_obj must be {}, got {} instead.'.format(Styler.__name__, type(styler_obj).__name__))

        if not isinstance(cols_to_style, (list, tuple, set, pd.Index)):
            cols_to_style = [cols_to_style]
        if not all(col in self.columns for col in cols_to_style):
            raise KeyError("one of the columns in {} wasn't found".format(cols_to_style))

        if overwrite_default_style:
            style_to_apply = styler_obj
        else:
            style_to_apply = Styler.combine(self._default_style, styler_obj)

        for col_name in cols_to_style:
            if style_header:
                self.columns[self.columns.get_loc(col_name)].style = style_to_apply
                self._has_custom_headers_style = True
            for index in self.index:
                if use_default_formats:
                    if isinstance(self.at[index, col_name].value, pd_timestamp):
                        style_to_apply.number_format = utils.number_formats.date_time
                    elif isinstance(self.at[index, col_name].value, dt.date):
                        style_to_apply.number_format = utils.number_formats.date
                    elif isinstance(self.at[index, col_name].value, dt.time):
                        style_to_apply.number_format = utils.number_formats.time_24_hours

                self.at[index, col_name].style = style_to_apply

        if width:
            self.set_column_width(columns=cols_to_style, width=width)

        return self

    def apply_headers_style(self, styler_obj, style_index_header=True, cols_to_style=None):
        """Apply style to the headers only
        :param Styler styler_obj: the styler object that contains the style to be applied
        :param bool style_index_header: if True then the style will also be applied to the header of the index column
        :param None|str|list|tuple|set cols_to_style: the columns to apply the style to, if not provided all the columns will be styled
        :return: self
        :rtype: StyleFrame
        """

        if not isinstance(styler_obj, Styler):
            raise TypeError('styler_obj must be {}, got {} instead.'.format(Styler.__name__, type(styler_obj).__name__))

        if cols_to_style is None:
            cols_to_style = self.data_df.columns
        if not isinstance(cols_to_style, (list, tuple, set, pd.Index)):
            cols_to_style = [cols_to_style]
        if not all(col in self.columns for col in cols_to_style):
            raise KeyError("one of the columns in {} wasn't found".format(cols_to_style))

        if style_index_header:
            self._index_header_style = styler_obj

        for column in cols_to_style:
            self.columns[self.columns.get_loc(column)].style = styler_obj
        self._has_custom_headers_style = True
        return self

    def set_column_width(self, columns, width):
        """Set the width of the given columns

        :param set|list|tuple columns: a single or a list/tuple/set of column name, index or letter to change their width
        :param int|float width: numeric positive value of the new width
        :return: self
        :rtype: StyleFrame
        """

        if not isinstance(columns, (set, list, tuple, pd.Index)):
            columns = [columns]
        try:
            width = float(width)
        except ValueError:
            raise TypeError('columns width must be numeric value')

        if width <= 0:
            raise ValueError('columns width must be positive')

        for column in columns:
            if not isinstance(column, (int, str_type, unicode_type, Container)):
                raise TypeError("column must be an index, column letter or column name")
            self._columns_width[column] = width

        return self

    def set_column_width_dict(self, col_width_dict):
        """
        :param dict col_width_dict: dictionary from tuple of columns to new width
        :return: self
        :rtype: StyleFrame
        """

        if not isinstance(col_width_dict, dict):
            raise TypeError("'col_width_dict' must be a dictionary")
        for cols, width in col_width_dict.items():
            self.set_column_width(cols, width)

        return self

    def set_row_height(self, rows, height):
        """ Set the height of the given rows

        :param int|list|tuple|set rows: a single row index or list, tuple or set of indexes to change their height
        :param height: numeric positive value of the new height
        :return: self
        :rtype: StyleFrame
        """

        if not isinstance(rows, (set, list, tuple, pd.Index)):
            rows = [rows]
        try:
            height = float(height)
        except ValueError:
            raise TypeError('rows height must be numeric value')

        if height <= 0:
            raise ValueError('rows height must be positive')
        for row in rows:
            try:
                row = int(row)
            except TypeError:
                raise TypeError("row must be an index")

            self._rows_height[row] = height

        return self

    def set_row_height_dict(self, row_height_dict):
        """
        :param dict row_height_dict: dictionary from tuple of rows to new height
        :return: self
        :rtype: StyleFrame
        """

        if not isinstance(row_height_dict, dict):
            raise TypeError("'row_height_dict' must be a dictionary")
        for rows, height in row_height_dict.items():
            self.set_row_height(rows, height)
        return self

    def rename(self, columns=None, inplace=False):
        """Renames the underlying dataframe's columns

        :param dict columns: a dictionary, old_col_name -> new_col_name
        :param inplace: whether to rename the columns inplace or return a new StyleFrame object
        :return: self if inplace=True, new StyleFrame object if inplace=False
        """

        if not isinstance(columns, dict):
            raise TypeError("'columns' must be a dictionary")

        sf = self if inplace else StyleFrame(self)

        new_columns = [col if col not in columns else Container(columns[col], col.style)
                       for col in sf.data_df.columns]

        sf._known_attrs['columns'] = sf.data_df.columns = new_columns

        sf._columns_width.update({new_col_name: sf._columns_width.pop(old_col_name)
                                  for old_col_name, new_col_name in columns.items()
                                  if old_col_name in sf._columns_width})
        return sf

    def style_alternate_rows(self, styles, **kwargs):
        """Applies the provided styles to rows in an alternating manner.

        :param list|tuple|set styles: styles to apply
        :return: self
        """

        num_of_styles = len(styles)
        split_indexes = (self.index[i::num_of_styles] for i in range(num_of_styles))
        for i, indexes in enumerate(split_indexes):
            self.apply_style_by_indexes(indexes, styles[i], **kwargs)
        return self

    def add_color_scale_conditional_formatting(self, start_type, start_value, start_color, end_type, end_value, end_color,
                                               mid_type=None, mid_value=None, mid_color=None, columns_range=None):
        """
        :param utils.conditional_formatting_types|str start_type: The type for the minimum bound
        :param start_value: The threshold for the minimum bound
        :param utils.colors|str start_color: The color for the minimum bound
        :param utils.conditional_formatting_types|str end_type: The type for the maximum bound
        :param end_value: The threshold for the maximum bound
        :param utils.colors|str end_color: The color for the maximum bound
        :param None|utils.conditional_formatting_types|str mid_type: The type for the middle bound
        :param mid_value: The threshold for the middle bound
        :param None|utils.colors|str mid_color: The color for the middle bound
        :param None|list|tuple columns_range: A two-elements list or tuple of columns to which the conditional formatting will be added
            to.
            If not provided at all the conditional formatting will be added to all columns.
            If a single element is provided then the conditional formatting will be added to the provided column.
            If two elements are provided then the conditional formatting will start in the first column and end in the second.
            The provided columns can be a column name, letter or index.
        :return: self
        """

        if columns_range is None:
            columns_range = (self.data_df.columns[0], self.data_df.columns[-1])

        if not isinstance(columns_range, (list, tuple)) or len(columns_range) not in (1, 2):
            raise TypeError("'columns_range' should be a list or a tuple with 1 or 2 elements")

        self._cond_formatting.append(ColorScaleConditionalFormatRule(start_type=start_type, start_value=start_value,
                                                                     start_color=start_color,
                                                                     mid_type=mid_type, mid_value=mid_value,
                                                                     mid_color=mid_color,
                                                                     end_type=end_type, end_value=end_value,
                                                                     end_color=end_color,
                                                                     columns_range=columns_range))

        return self
